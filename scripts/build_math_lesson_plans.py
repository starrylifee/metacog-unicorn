from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / 'source'
OUTPUT_PATH = ROOT / 'src' / 'data' / 'mathLessonPlans.json'
PAGE_ONLY_RE = re.compile(r'^\d+(?:~\d+)?$')
GRADE_SEMESTER_RE = re.compile(r'(\d)학년\s*(\d)학기')


def normalize_text(value: object) -> str:
    if value is None:
        return ''

    text = str(value).replace('\u3000', ' ')
    text = text.replace('\r', '\n')
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n+', '\n', text)
    return text.strip()


def clean_unit_name(value: object) -> str:
    text = normalize_text(value).replace('\n', ' ')
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'^\d+\.\s*', '', text)
    return text.strip()


def append_lesson(units: OrderedDict[str, list[str]], unit_name: str, lesson: str) -> None:
    lesson = normalize_text(lesson).replace('\n', ' ')
    lesson = re.sub(r'\s+', ' ', lesson).strip()

    if not unit_name or not lesson or PAGE_ONLY_RE.fullmatch(lesson):
        return

    unit_lessons = units.setdefault(unit_name, [])
    if lesson not in unit_lessons:
        unit_lessons.append(lesson)


def parse_lesson_lines(raw_cells: list[object]) -> tuple[list[str], bool]:
    lines: list[str] = []

    for cell in raw_cells:
        cell_text = normalize_text(cell)
        if not cell_text:
            continue
        for line in cell_text.split('\n'):
            clean_line = normalize_text(line)
            if clean_line:
                lines.append(clean_line)

    if not lines:
        return [], False

    continuation_only = all(line.startswith('-') for line in lines)
    titles: list[str] = []
    current: str | None = None

    for line in lines:
        if line.startswith('•'):
            if current:
                titles.append(current.strip())
            current = normalize_text(line.lstrip('•').strip())
            continue

        if line.startswith('-'):
            suffix = normalize_text(line[1:].strip())
            if current:
                current = f'{current} -{suffix}'
            elif titles:
                titles[-1] = f'{titles[-1]} -{suffix}'
            else:
                current = f'-{suffix}'
            continue

        if current:
            current = f'{current} {line}'.strip()
        elif titles:
            titles[-1] = f'{titles[-1]} {line}'.strip()
        else:
            current = line

    if current:
        titles.append(current.strip())

    return titles, continuation_only


def detect_title_start(table: list[list[object]]) -> int:
    header_row = next(
        (
            row
            for row in table[:3]
            if any(normalize_text(cell) == '차시명' for cell in row)
        ),
        None,
    )
    if header_row:
        return next(idx for idx, cell in enumerate(header_row) if normalize_text(cell) == '차시명')

    first_width = len(table[0]) if table else 0
    if first_width >= 11:
        return 10
    if first_width >= 7:
        return first_width - 1
    return first_width


def parse_pdf_file(path: Path) -> tuple[str, str, OrderedDict[str, list[str]]]:
    match = GRADE_SEMESTER_RE.search(path.stem)
    if not match:
        raise ValueError(f'Could not infer grade/semester from {path.name}')

    grade, semester = match.groups()
    units: OrderedDict[str, list[str]] = OrderedDict()
    current_unit = ''

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table:
                    continue

                title_start = detect_title_start(table)
                if title_start >= len(table[0]):
                    continue

                for row in table:
                    if not row:
                        continue

                    row_text = ' '.join(normalize_text(cell) for cell in row if normalize_text(cell))
                    if '차시명' in row_text:
                        continue

                    unit_name = clean_unit_name(row[2] if len(row) > 2 else '')
                    if unit_name:
                        current_unit = unit_name
                        units.setdefault(current_unit, [])

                    title_cells = list(row[title_start:]) if len(row) > title_start else []
                    titles, continuation_only = parse_lesson_lines(title_cells)

                    if not current_unit or not titles:
                        continue

                    if continuation_only:
                        suffix = titles[0].lstrip('-').strip()
                        if suffix and units[current_unit]:
                            existing_lesson = units[current_unit][-1]
                            if not existing_lesson.endswith(f" -{suffix}"):
                                units[current_unit][-1] = f"{existing_lesson} -{suffix}"
                        continue

                    for title in titles:
                        append_lesson(units, current_unit, title.lstrip('-').strip())

    return grade, semester, units


def parse_xlsx_file(path: Path) -> tuple[str, str, OrderedDict[str, list[str]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    headers = [normalize_text(value) for value in next(rows)]
    index = {header: idx for idx, header in enumerate(headers)}

    grade_idx = index['학년']
    semester_idx = index['학기']
    subject_idx = index['편제']
    unit_idx = index['단원']
    lesson_idx = index['학습내용']

    units: OrderedDict[str, list[str]] = OrderedDict()
    grade = ''
    semester = ''

    for row in rows:
        subject = normalize_text(row[subject_idx])
        if subject != '수학':
            continue

        grade = re.sub(r'\D', '', normalize_text(row[grade_idx])) or grade
        semester = re.sub(r'\D', '', normalize_text(row[semester_idx])) or semester
        unit_name = clean_unit_name(row[unit_idx])
        lesson_title = normalize_text(row[lesson_idx])
        append_lesson(units, unit_name, lesson_title)

    if not grade or not semester:
        raise ValueError(f'Could not infer grade/semester from {path.name}')

    return grade, semester, units


def merge_section(
    target: dict[str, dict[str, list[dict[str, list[str]]]]],
    grade: str,
    semester: str,
    units: OrderedDict[str, list[str]],
) -> None:
    target.setdefault(grade, {})
    target[grade][semester] = [
        {'unit': unit_name, 'lessons': lessons}
        for unit_name, lessons in units.items()
        if unit_name and lessons
    ]


def build_dataset() -> dict[str, dict[str, dict[str, list[dict[str, list[str]]]]]]:
    grades: dict[str, dict[str, list[dict[str, list[str]]]]] = {}

    for pdf_path in sorted(SOURCE_DIR.glob('*.pdf')):
        grade, semester, units = parse_pdf_file(pdf_path)
        merge_section(grades, grade, semester, units)

    for xlsx_path in sorted(SOURCE_DIR.glob('*.xlsx')):
        grade, semester, units = parse_xlsx_file(xlsx_path)
        merge_section(grades, grade, semester, units)

    return {'grades': grades}


def main() -> None:
    data = build_dataset()
    OUTPUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

    for grade in sorted(data['grades'].keys(), key=int):
        semesters = data['grades'][grade]
        unit_count = sum(len(units) for units in semesters.values())
        print(f"grade {grade}: semesters={sorted(semesters.keys())} units={unit_count}")


if __name__ == '__main__':
    main()
